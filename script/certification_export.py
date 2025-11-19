import pandas as pd
import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side,Alignment
import calendar
import locale
from .db_config import get_config


def export_monthly_certification(user,password,schema,output_path,date):

    try:

        locale.setlocale(locale.LC_TIME, 'es_ES.utf8')


        # Crear un archivo de Excel vacío
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # Definir el color de fondo y la fuente en negrita
        summary_fill = PatternFill(start_color='B7D6FF', end_color='B7D6FF', fill_type='solid')  # Color de fondo azul claro
        bold_font = Font(bold=True)
        pink_font = Font(bold=True, color="FF1493")# Fuente en negrita

        # Crear un nuevo archivo Excel
        wb = Workbook()


        #Conexión con BBDD usando configuración centralizada
        config = get_config()
        conexion = mysql.connector.connect(
                    host=config.host,
                    port=config.port,
                    user=user,
                    password=password
                )

        # datos presupuesto
        sql_query = "SELECT * FROM " + schema + ".vw_presupuesto"
        df_budget = pd.read_sql(sql_query, conexion)

        #datos certificaciones
        year = date.split('-')[0]
        month = date.split('-')[1]
        sql_query = f"SELECT * FROM {schema}.vw_certificaciones WHERE YEAR(fecha_certificacion) = {year} AND MONTH(fecha_certificacion) = {month}"
        df_cost = pd.read_sql(sql_query, conexion)

        # consultamos los datos economicos del proyecto
        sql_query = f"SELECT * FROM manager.tbl_proyectos WHERE codigo = '{schema}'"
        df_project = pd.read_sql(sql_query, conexion)
        sql_query = f"SELECT * FROM manager.tbl_proy_presupuesto WHERE id_proyecto = '{df_project.id.values[0]}'"
        df_project_economic = pd.read_sql(sql_query, conexion)

        gg=df_project_economic.gastos_generales.values[0]
        bi=df_project_economic.beneficio_industrial.values[0]
        tax=df_project_economic.iva.values[0]

        #recogemos registros seleccionados
        register_select = df_cost.arqueta.unique()

        # Iterar por cada registro en `register_select`
        for item in register_select:
            # Filtrar los DataFrames
            df_filter_budget = df_budget[df_budget.arqueta == item]
            df_filter_cost = df_cost[df_cost.arqueta == item]

            # Combinar datos
            df_merge = pd.merge(
                df_filter_cost[['cod_proyecto', 'arqueta', 'cod_partida', 'naturaleza','unidad','resumen', 'descripcion', 'precio_unitario',
                                'cantidad_certificada', 'coste_total', 'cod_capitulo', 'capitulo', 'fecha_certificacion']],
                df_filter_budget[['cod_proyecto', 'arqueta', 'cod_partida', 'naturaleza','unidad', 'resumen', 'descripcion', 'precio_unitario',
                                  'cantidad', 'coste_total', 'cod_capitulo', 'capitulo']],
                on=['cod_proyecto', 'arqueta', 'cod_partida','naturaleza','unidad', 'resumen', 'descripcion', 'precio_unitario', 'cod_capitulo',
                    'capitulo'],
                how='outer'
            )

            # Crear una nueva hoja para cada registro
            sheet = wb.create_sheet(title=str(item))

            head = ['Código', 'naturaleza','unidad','Resumen', 'Precio Unitario',
                               'Cantidad Presupuesto', 'Coste Presupuesto',
                               'Cantidad Certificada', 'Coste Certificación']

            # Escribir los datos en la siguiente fila y ponerlos en negrita
            for col, value in enumerate(head, start=1):
                cell = sheet.cell(row=1, column=col, value=value)
                # Aplicar negrita al texto de la celda
                cell.font = Font(bold=True)

            chapters = df_merge.cod_capitulo.unique()
            # Añadir resumen al inicio de la hoja
            for chp in chapters:
                sheet_data = []
                df_filter = df_merge[df_merge.cod_capitulo == chp]
                df_filter = df_filter.fillna(0)
                summary = df_filter.groupby(['cod_capitulo', 'capitulo']).agg(
                    {'coste_total_x': 'sum', 'coste_total_y': 'sum'}).reset_index()

                # Generar la linea del capitualo
                chapter_data =[summary['cod_capitulo'].iloc[0],'Capitulo','',summary['capitulo'].iloc[0],'','',summary['coste_total_y'].iloc[0],'',summary['coste_total_x'].iloc[0]]
                last_row = sheet.max_row
                for col, value in enumerate(chapter_data, start=1):
                    cell = sheet.cell(row=last_row+1, column=col, value=value)
                    # Aplicar negrita al texto de la celda
                    cell.font = Font(bold=True)
                    cell.fill = summary_fill

                # Añadir los datos detallados debajo del resumen (por cada item del capítulo)
                for index, row in df_filter.iterrows():
                    code = row['cod_partida']
                    type = row['naturaleza']
                    unit = row['unidad']
                    resume = row['resumen']
                    description = row['descripcion']
                    price = row['precio_unitario']
                    amount_budget = row['cantidad']
                    cost_budget = row['coste_total_y']
                    amount_cost = row['cantidad_certificada']
                    cost_cost = row['coste_total_x']

                    sheet_data=[code, type, unit, resume, price, amount_budget, cost_budget, amount_cost, cost_cost]

                    last_row = sheet.max_row
                    for col, value in enumerate(sheet_data, start=1):
                        cell = sheet.cell(row=last_row + 1, column=col, value=value)

                    sheet_data = ['', '', '', description, '', '', '', '', '']

                    last_row = sheet.max_row
                    for col, value in enumerate(sheet_data, start=1):
                        cell = sheet.cell(row=last_row + 1, column=col, value=value)


                last_row = sheet.max_row
                for i in range(last_row):
                    cell = sheet.cell(row=i + 1, column=7)
                    cell.font = pink_font
                    cell = sheet.cell(row=i + 1, column=9)
                    cell.font = pink_font

            #totales presupuesto
            pem_budget = df_merge.coste_total_x.sum()
            gg_budget = pem_budget * gg / 100
            bi_budget = pem_budget * bi / 100
            total_budget = pem_budget + gg_budget + bi_budget

            #totaltes certificaciones
            pem_cost = df_merge.coste_total_y.sum()
            gg_cost = pem_cost * gg / 100
            bi_cost = pem_cost * bi / 100
            total_cost = pem_cost + gg_cost + bi_cost

            pem_data=['','','','PRESUPUESTO DE EJECUCIÓN MATERIAL','','',pem_budget,'',pem_cost]
            last_row = sheet.max_row
            for col, value in enumerate(pem_data, start=1):
                cell = sheet.cell(row=last_row + 1, column=col, value=value)
                cell.font = Font(bold=True)
                cell.border = Border(bottom=Side(border_style="thin", color="000000"))

            gg_data = ['', '', '',f'{round(gg,2)}% Gastos generales', '', '', gg_budget, '', gg_cost]
            last_row = sheet.max_row
            for col, value in enumerate(gg_data, start=1):
                cell = sheet.cell(row=last_row + 1, column=col, value=value)


            bi_data = ['', '', '', f'{round(bi,2)}% Beneficio industrial', '', '', bi_budget, '', bi_cost]
            last_row = sheet.max_row
            for col, value in enumerate(bi_data, start=1):
                cell = sheet.cell(row=last_row + 1, column=col, value=value)
                cell.border = Border(bottom=Side(border_style="thin", color="000000"))

            total_data = ['', '', '', f"TOTAL {item} ", '', '', total_budget, '', total_cost]
            last_row = sheet.max_row
            for col, value in enumerate(total_data, start=1):
                cell = sheet.cell(row=last_row + 1, column=col, value=value)
                cell.font = Font(bold=True)


        resume_budget = df_budget[df_budget['arqueta'].isin(register_select)].groupby('arqueta').agg({'coste_total':'sum'}).reset_index()
        resume_cost = df_cost.groupby('arqueta').agg({'coste_total':'sum'}).reset_index()
        resume = pd.merge(resume_budget,resume_cost, on='arqueta', how='outer')



        sheet = wb.create_sheet(title='Resumen')
        month_name = calendar.month_name[int(month)].upper()
        cell = sheet.cell(row=1, column=1, value=f"RESUMEN CERTIFICACIÓN {month_name}")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=16, bold=True)
        cell.fill = summary_fill
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=4
        )

        cell = sheet.cell(row=1, column=5, value=f"EJECUTADO")
        cell.font = Font(color="FF0000", underline="single", bold=True, size=14)
        sheet.merge_cells(
            start_row=1,
            start_column=5,
            end_row=2,
            end_column=5
        )

        sheet.insert_rows(3)

        # Iterar por cada registro en `register_select`
        for item in register_select:
            sql_query = f"SELECT * FROM {schema}.tbl_inventario WHERE codigo = '{item}'"
            df_register = pd.read_sql(sql_query, conexion)
            sql_query = f"SELECT * FROM {schema}.tbl_municipios WHERE id = '{df_register.id_municipio.values[0]}'"
            df_locality = pd.read_sql(sql_query, conexion)
            title = item + " " +df_locality.NAMEUNIT.values[0]
            # Filtrar los DataFrames
            total_item_budget = df_budget[df_budget.arqueta == item].coste_total.sum()
            total_item_cost = df_cost[df_cost.arqueta == item].coste_total.sum()
            diff = total_item_budget - total_item_cost
            pct = (total_item_cost *100)/total_item_budget

            sheet_data = [title, '', '', total_item_budget, total_item_cost , diff, pct]

            last_row = sheet.max_row
            for col, value in enumerate(sheet_data, start=1):
                cell = sheet.cell(row=last_row + 1, column=col, value=value)

        last_row = sheet.max_row
        sheet.insert_rows(last_row+1)

        sum_budget = df_budget[df_budget.arqueta.isin(register_select)].coste_total.sum()
        sum_cost = df_cost.coste_total.sum()
        diff = df_budget[df_budget.arqueta.isin(register_select)].coste_total.sum()- df_cost.coste_total.sum()
        pct = (df_cost.coste_total.sum()* 100) / df_budget[df_budget.arqueta.isin(register_select)].coste_total.sum()

        sheet_data = ['CERTIFICACIÓN DE EJECUCIÓN MATERIAL', '', '', sum_budget, sum_cost , diff, pct]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = summary_fill

        last_row = sheet.max_row
        sheet.insert_rows(last_row+1)

        gg_budget = sum_budget * gg / 100
        bi_budget = sum_budget * bi / 100
        total_budget = sum_budget + gg_budget + bi_budget
        tax_budget = tax * total_budget
        total_budget_tax = total_budget +  tax_budget

        gg_cost = sum_cost * gg / 100
        bi_cost = sum_cost * bi / 100
        total_cost = sum_cost + gg_cost + bi_cost
        tax_cost = tax * total_cost
        total_cost_tax = total_cost +  tax_cost


        sheet_data = ['Gastos generales', f"{round(gg)} %", '', gg_budget, gg_cost]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)

        sheet_data = ['Beneficio industrial', f"{round(bi)} %", '', bi_budget, bi_cost]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)

        last_row = sheet.max_row
        sheet.insert_rows(last_row+1)

        sheet_data = ['CERTIFICACIÓN EJECUCIÓN POR CONTRATA', '', '', total_budget, total_cost]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = summary_fill

        last_row = sheet.max_row
        sheet.insert_rows(last_row+1)

        sheet_data = ['IVA', f"{round(tax)} %", '', tax_budget, tax_cost]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)

        last_row = sheet.max_row
        sheet.insert_rows(last_row+1)

        diff = total_budget_tax-total_cost_tax
        sheet_data = [f'CERTIFICACIÓN MES DE {month_name}', '', '', total_budget_tax, total_cost_tax,diff ]

        last_row = sheet.max_row
        for col, value in enumerate(sheet_data, start=1):
            cell = sheet.cell(row=last_row + 1, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = summary_fill

        # Aplicar formato a los números negativos
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=6, max_col=6):  # Iterar por la columna 1
            for cell in row:
                if cell.value is None:  # Si la celda está vacía
                    cell.value = ""
                else:  # Si el valor es negativo
                    cell.font = Font(color="FF0000")

        del wb['Sheet']
        wb.save(output_path)

        return "ok"

    except Error as e:
        print(f"Error: {e}")
        return e

